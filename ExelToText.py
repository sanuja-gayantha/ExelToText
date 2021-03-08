from datetime import datetime as dt
import os
import openpyxl
import ctypes
import pandas as pd

pd.set_option('display.max_columns', 15000) # number of columns to be displayed
pd.set_option('display.max_rows', 15000)   # max table width to display


currentPath = os.getcwd()
folder_full_path = os.path.join(currentPath, "Exel_folder")
if not os.path.exists(folder_full_path):
    os.makedirs(folder_full_path)  
file_names = os.listdir(folder_full_path)



def acc_number_length(Account_number):
    account_data_dict = {12:"", 11:"0", 10:"00", 9:"000", 8:"0000", 7:"00000", 6:"000000", 5:"0000000", 4:"00000000", 3:"000000000", 2:"0000000000", 1:"00000000000", 0:"000000000000"}
    Account_number_length = len(Account_number)
    if Account_number_length in account_data_dict:
        Correct_account_number = account_data_dict[Account_number_length] + Account_number
        
        return Correct_account_number

def cor_name(Name):
    LName = list(filter(None,Name))
    if (LName[-1] == " "):
        Name = Name[:-1]
        return Name
    else:
        return Name

def fin_name(Correct_name):
    Name_length = len(Correct_name)
    if(Name_length > 20):
        if_car_len = Name_length - 20
        f_name = Correct_name[:-if_car_len]
        return f_name

    else:
        name_data_dict = {0:"", 1:" ", 2:"  ", 3:"   ", 4:"    ", 5:"     ", 6:"      ", 7:"       ", 8:"        ", 9:"         ", 10:"          ", 11:"           ", 12:"            ", 
                          13:"             ", 14:"              ", 15:"               ", 16:"                ", 17:"                 ", 18:"                  ", 19:"                   ",
                          20:"                    "}

        else_car_len = 20 - Name_length
        if else_car_len in name_data_dict:
            f_name = Correct_name + name_data_dict[else_car_len] 
            return f_name

def cor_amount(Amount):
    Samount = str(Amount).split(".")
    if(len(Samount) > 1):
        Camount = Samount[0] + Samount[1]
    else:
        Camount = Samount[0] + "00"
    CCAmount = acc_number_length(str(Camount))

    return CCAmount

def company_name(file_name):
    if("X" == file_name.split(" ")[0]):
        company = "X company"
    elif("Y"):
        company = "Y company"
    else:
        pass
    return company

def cor_empNO(Number):
    emp_num = Number.lstrip("0")
    Name_length = len(emp_num)

    name_data_dict = {0:"", 1:" ", 2:"  ", 3:"   ", 4:"    ", 5:"     ", 6:"      ", 7:"       ", 8:"        ", 9:"         ", 10:"          ", 11:"           ", 12:"            ", 
                          13:"             ", 14:"              ", 15:"               "}

    else_car_len = 15 - Name_length
    if else_car_len in name_data_dict:
        f_name = emp_num + name_data_dict[else_car_len] 
        return f_name   

def sec_arg(file_name):
    if("INC" == file_name.split(" ")[1].upper()):
        sec_argu = "INC"
    elif ("BONUS" == file_name.split(" ")[1].upper()):
        sec_argu = "BONUS"
    elif("SALARY" == file_name.split(" ")[1].upper()):
        sec_argu = "SALARY"
    else:
        pass

    Name_length = len(sec_argu)
    name_data_dict = {0:"", 1:" ", 2:"  ", 3:"   ", 4:"    ", 5:"     ", 6:"      ", 7:"       ", 8:"        ", 9:"         ", 10:"          ", 11:"           ", 12:"            ", 
                          13:"             ", 14:"              ", 15:"               "}

    else_car_len = 15 - Name_length
    if else_car_len in name_data_dict:
        f_name = sec_argu + name_data_dict[else_car_len] 
        return f_name   

def final_line(company, sec_argument, current_date, Total_amount):

    Samount = str(Total_amount).split(".")
    if(len(Samount) > 1 and len(str(Samount[1])) > 1):
        Camount = Samount[0] + Samount[1]
    elif(len(Samount) > 1 and len(str(Samount[1])) == 1):
        Camount = Samount[0] + "00"
    else:
        print("Something went wrong in final_line 1 !!")

    TTamount = acc_number_length(str(Camount))

    if("X company" == company):
        fprint_line = ("0000"+ "A1"+ "B1"+ "C1"+ "X company "+ "D1"+ "00"+ "1"+ 
            "000000"+  str(TTamount)+ "SLR"+ "A1"+ "B1"+ "C1"+ "X company"+ "                "+
            str(sec_argument)+  str(current_date)+ "      "+ "@")
    elif("Y company" == company):
        fprint_line = ("0000"+ "A2"+ "B2"+ "C2"+ "Y company     "+ " D2"+ "00"+ "1"+ 
            "000000"+  str(TTamount)+ "SLR"+ "A2"+ "B2"+ "C2"+ "Y company"+ "                     "+
            str(sec_argument)+  str(current_date)+ "      "+ "@")
    else:
        print("Somethig went wrong final_line 2 !!")

    return fprint_line


def CC_final_account_number(Correct_account_number):
    C_len = len(Correct_account_number)
    if(C_len > 12):
        number_of_car_remove = C_len - 12
        CCC_number = Correct_account_number[number_of_car_remove:]
    else:
        CCC_number = Correct_account_number

    return CCC_number

def EEBranch_code(Branch_code):
    B_len = len(Branch_code)
    B_data_dict = {1:"00", 2:"0", 3:""}
    if B_len in B_data_dict:
        Correct_branch_number = B_data_dict[B_len] + Branch_code
        
        return Correct_branch_number



def main():   
    # MAIN PART OF THE SCRIPT
        
    if(file_names):
        current_date = dt.today().strftime('%y%m%d')
        

        for file_name in file_names:
            if file_name.endswith('.xlsx'):
                # wcom_name = os.path.join(folder_full_path, file_name)
                wb = openpyxl.load_workbook(os.path.join(folder_full_path, file_name))
                sheet = wb[wb.sheetnames[0]]

                company = company_name(file_name)
                sec_argument = sec_arg(file_name)

                text_file_name =  file_name.split(".")[0]+ ".txt"
                # print(text_file_name) 
                tcom_name = os.path.join(folder_full_path, text_file_name)          
                text_file = open(tcom_name, "w")
                Total_amount = 0.00

                if("X" == company):
                    CBankC = "A1"
                    CBranchC = "B1"
                    CAnum =  "C1"
                    Cspaces = " "
                elif("Y" == company):
                    CBankC = "A2"
                    CBranchC = "B2"
                    CAnum =  "C2"
                    Cspaces = "      "
                else:
                    pass 

                for i in range(2,sheet.max_row):
                    Id = sheet.cell(row=i,column=1).value
                    Number = str(sheet.cell(row=i,column=2).value)
                    emp_no = cor_empNO(str(Number))

                    Name = sheet.cell(row=i,column=3).value.upper()
                    final_name = fin_name(cor_name(Name))

        
                    Bank = str(sheet.cell(row=i,column=4).value)
                    Bank_code = str(sheet.cell(row=i,column=5).value)    
                    Branch = str(sheet.cell(row=i,column=6).value)
                    Branch_code = str(sheet.cell(row=i,column=7).value)
                    EBranch_code = EEBranch_code(Branch_code)

                    Account_number = str(sheet.cell(row=i,column=8).value)
                    CC_account_number = CC_final_account_number(Account_number)
                    Correct_account_number = acc_number_length(str(CC_account_number))
                    

                    Amount = sheet.cell(row=i,column=9).value
                    Correct_amount = cor_amount(round(Amount, 2))
                    Total_amount += round(Amount, 3)
                    # print(Id,Amount,Total_amount)
                    

                    pint_line = "0000"+ str(Bank_code)+ str(EBranch_code)+ str(Correct_account_number)+ str(final_name)+ "23"+ "00"+ "0"+ "000000"+  str(Correct_amount)+ "SLR"+ str(CBankC)+ str(CBranchC)+ str(CAnum)+ str(company)+ Cspaces+ str(emp_no)+ str(sec_argument)+  str(current_date)+ "      "+ "@"  
                    text_file.write(pint_line+ "\n")


                fprint_line = final_line(company, sec_argument, current_date, round(Total_amount,2))
                text_file.write(fprint_line+ "\n")             
                text_file.close()






if __name__ == '__main__':

    try:
        main()
        ctypes.windll.user32.MessageBoxW(0, "         Done.", "Message", 0)

    except Exception as e:
        ctypes.windll.user32.MessageBoxW(0, "         Something went wrong", "Message", 0)
        pass



























